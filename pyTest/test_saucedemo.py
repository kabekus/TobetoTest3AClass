from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium import webdriver
import pytest
import openpyxl

class Test_Saucedemo:
    def setup_method(self):
        #Her test başlangıcında çalışacak fonksiyon
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument("--headless")
        self.driver = webdriver.Chrome(options=chrome_options)
        self.driver.get("https://www.saucedemo.com/") 
    

    def teardown_method(self):
        #Her test bitiminde çalışır   
        self.driver.quit()

    def getData(): # Bir anatasyon içinde çağıracağım bir fonksiyon varsa ona 'self' parametresi verilmez
        excelFile = openpyxl.load_workbook("data/invalidLogin.xlsx") 
        sheet = excelFile["Sheet1"]
        rows = sheet.max_row
        dataInExcel = []

        for i in range(2,rows+1):
            userNameXL = sheet.cell(i,1).value 
            passwordXL = sheet.cell(i,2).value
            dataInExcel.append((userNameXL,passwordXL))

        return dataInExcel
        # return [("abc","123"),("1","1"),("deneme","secret")]

    @pytest.mark.parametrize("userName,password",getData())
    def test_invalid_login(self,userName,password):
        # sleep(30)   #verilen süre kadar açılan ekranı durdurmak için kullandık
        userNameInput = WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,"user-name")))
        userNameInput.send_keys(userName)
        passwordInput = WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,"password")))
        passwordInput.send_keys(password)
        loginButton = WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,"login-button")))
        loginButton.click()
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"
       
    @pytest.mark.skip
    def test_valid_login(self):
        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,"user-name")))
        username = self.driver.find_element(By.ID,"user-name")
        
        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,"password")))
        password = self.driver.find_element(By.NAME,"password")

        WebDriverWait(self.driver,5).until(expected_conditions.visibility_of_element_located((By.ID,"login-button")))
        loginButton = self.driver.find_element(By.ID,"login-button")
        loginButton.click()
        appLogo = self.driver.find_element(By.CLASS_NAME,"app_logo")
        assert appLogo.text == "Swag Labs"
        
        
